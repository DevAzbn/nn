
from openpyxl import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

# ws2 = wb.create_sheet("Mysheet", 0) #-1 2

ws.title = 'New Title'
ws.sheet_properties.tabColor = "000000"

# ws3 = wb["New Title"]

print(wb.sheetnames)
for sheet in wb:
	print(sheet.title)

# source = wb.active
# target = wb.copy_worksheet(source)

c = ws['A4']
ws['A4'] = 4

wb.save('wb.xlsx')

# d = ws.cell(row = 4, column = 2, value = 10)

# for x in range(1, 101):
# 	for y in range(1,101):
# 		ws.cell(row=x, column=y)

# cell_range = ws['A1':'C2']

# colC = ws['C']
# col_range = ws['C:D']
# row10 = ws[10]
# row_range = ws[5:10]

# for row in ws.iter_rows(min_row = 1, max_col = 3, max_row = 2):
# 	for cell in row:
# 		print(cell)

# for col in ws.iter_cols(min_row = 1, max_col = 3, max_row = 2):
# 	for cell in col:
# 		print(cell)

# tuple(ws.rows)
# tuple(ws.columns)

# for row in ws.values:
# 	for value in row:
# 		print(value)

# for row in ws.iter_rows(min_row = 1, max_col = 3, max_row = 2, values_only = True):
# 	print(row)

# c.value = 'hello, world'
# print(c.value)

# wb.save('balances.xlsx')

# from openpyxl import load_workbook
# wb2 = load_workbook('test.xlsx')
# print wb2.sheetnames

# from tempfile import NamedTemporaryFile
# from openpyxl import Workbook
# wb = Workbook()
# with NamedTemporaryFile() as tmp:
# 	wb.save(tmp.name)
# 	tmp.seek(0)
# 	stream = tmp.read()

# wb = load_workbook('document.xlsx')
# wb.template = True
# wb.save('document_template.xltx')

# wb = load_workbook('document_template.xltx')
# wb.template = False
# wb.save('document.xlsx', as_template=False)